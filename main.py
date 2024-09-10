from openpyxl.cell import cell
import qrcode
import openpyxl
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("product table.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(3, dataframe1.max_row):
    for col in range(1, dataframe1.max_column,3):
        cell = dataframe1.cell(row,col)
        if(cell.value !=None):
            key = cell.value
            value = dataframe1.cell(row,col+1).value
            typeProduct = dataframe1.cell(1,col).value
            col=col+1
            print(value)
            img= qrcode.make("reference name: "+key +" and product Name :"+ value+" type:"+typeProduct)
            type(img)
            img.save("./qrCode/"+value+".png")
